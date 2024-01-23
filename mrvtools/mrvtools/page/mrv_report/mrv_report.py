
import frappe,json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from frappe.utils import get_site_base_path,now


monitoringYears = []

@frappe.whitelist()
def get_project_details(project = None):
	columns, data = get_project_columns(project), get_project_datas(project)
	return columns, data
@frappe.whitelist()
def get_mitigation_details(project = None):
	columns, data = get_mitigation_columns(project), get_mitigation_datas(project)
	return columns, data
@frappe.whitelist()
def get_adaptation_details(project = None):
	columns, data = get_adaptation_columns(project), get_adaptation_datas(project)
	return columns, data
@frappe.whitelist()
def get_sdg_details(project = None):
	columns, data = get_sdg_columns(project), get_sdg_datas(project)
	return columns, data
@frappe.whitelist()
def get_finance_details(project = None):
	columns, data = get_finance_columns(project), get_finance_datas(project)
	return columns, data



def get_project_columns(project = None):
	col = []
	if project:	
		
		col.append("Project Title" + ":Link/Project")
		col.append("Cost in USD" + ":Float")
		col.append("Location" + ":Data")
		col.append("Implementing entity or entities" + ":Data")
		col.append("Other Agency" + ":Data")
		col.append("Start Date" + ":Data")
		col.append("Lifetime in Years" + ":Int")
	keys_only = [item.split(':')[0] for item in col]
	return keys_only
def get_mitigation_columns(project = None):
	col = []
	if project:	
		
		col.append("Performance Indicator" + ":Data")
		col.append("Unit" + ":Data")
		col.append("Expected Value" + ":Data")
		totalMonitoringYearsMitigation = frappe.db.get_all('Mitigation Monitoring Information',
					filters={'proj_id':project},
					fields = ['monitoring_year'],
					order_by = 'monitoring_year asc' )
		monitoringYears = totalMonitoringYearsMitigation
		for i in totalMonitoringYearsMitigation:
			col.append(f"{i.monitoring_year}" + ":Data")
	keys_only = [item.split(':')[0] for item in col]
	return keys_only

def get_adaptation_columns(project = None):
	col = []
	if project:	
		
		col.append("Category" + ":Data")
		col.append("Question" + ":Data")
		col.append("Expected Value" + ":Float")
		totalMonitoringYearsAdaptation = frappe.db.get_all('Adaptation Monitoring Information',
						filters={'proj_id':project},
						fields = ['monitoring_year'],
						order_by = 'monitoring_year asc' )
		for i in totalMonitoringYearsAdaptation:
			col.append(f"{i.monitoring_year}" + ":Data")
	keys_only = [item.split(':')[0] for item in col]
	return keys_only

def get_sdg_columns(project = None):
	col = []
	if project:	
		
		
		col.append("Category" + ":Data")
		col.append("Question" + ":Data")
		col.append("Expected Value" + ":Float")
		totalMonitoringYearsAdaptation = frappe.db.get_all('SDG Monitoring Information',
						filters={'proj_id':project},
						fields = ['monitoring_year'],
						order_by = 'monitoring_year asc' )
		for i in totalMonitoringYearsAdaptation:
			col.append(f"{i.monitoring_year}" + ":Data")
	keys_only = [item.split(':')[0] for item in col]
	return keys_only

def get_finance_columns(project = None):
	col = []
	if project:
		
		col.append("Finance Data" + ":Data")
		name = frappe.db.get_value('Climate Finance', {'project_id': f'{project}'}, 'name')
		totalMonitoringYearsFinance = frappe.db.get_all(
			"Climate Finance Disbursement Schedule ChildTable",
			fields=['financial_year'],
			filters = {"parent" : name},
			order_by = 'financial_year')
		frappe.log_error("datataa",totalMonitoringYearsFinance)
		for i in totalMonitoringYearsFinance:
			col.append(f"{i.financial_year}" + ":Int")
		last_doc = frappe.get_last_doc("Climate Finance Monitoring Information",{'project_id': f'{name}'})
		totalMonitoringYears = frappe.db.get_all(
			"Climate Finance Total Budget Disbursement ChildTable",
			fields=['financial_year'],
			filters = {"parent" : last_doc.name},
			group_by = 'financial_year' )
		frappe.log_error("dataFinance",totalMonitoringYears)
		for i in totalMonitoringYears:
			if f"{i.financial_year}" + ":Int" not in col:
				col.append(f"{i.financial_year}" + ":Int")
	keys_only = [item.split(':')[0] for item in col]
	return keys_only

def get_project_datas(project = None):
	conditions = ""
	if project:
		
		conditions += f" AND P.name like '{project}'"
		query= f"""
			SELECT
				P.project_name as project_title,
				P.costusd as cost_in_usd,
				P.location,
				P.implementing_entity as implementing_entity_or_entities,
				P.other_agency,
				P.start_date,
				P.lifetime as lifetime_in_years,
				P.name
			FROM
				`tabProject` as P
			WHERE 
				P.docstatus != 2
				{conditions}
			"""
		result = frappe.db.sql(query, as_dict=1)
		values_only = [list({k: v for k, v in item.items() if k != 'name'}.values()) for item in result]
		return values_only
	
@frappe.whitelist(allow_guest=True)
def get_mitigation_datas(project = None):
	conditions = ""
	if project:
		
		conditions += f" AND MI.project_id = '{project}'"

		expectedValue = frappe.db.get_value('Mitigations',{"project_id":f'{project}'},'expected_annual_ghg')
		output = [{"Performance Indicator":"GHG Emissions Achieved (tCO2e)","Unit":"tCO2e","Expected Value":expectedValue}]
		monitoringYears = frappe.db.get_all('Mitigation Monitoring Information',
				filters={'proj_id':project},
				fields = ['monitoring_year'],
				order_by = 'monitoring_year asc' )
		for i in monitoringYears:
			currentValue = frappe.db.get_value('Mitigation Monitoring Information',{"proj_id":f'{project}',"monitoring_year":i.monitoring_year},'actual_annual_ghg')
			output[0][f'{i.monitoring_year}'] = currentValue if currentValue else 0
		
		query= f"""
			SELECT
				MCT.performance_indicator,
				MCT.unit,
				MCT.expected_value
			FROM
				`tabMitigations` as MI
			INNER JOIN
				`tabMitigation Performance Indicator ChildTable` as MCT
			ON
				MI.name = MCT.parent
			WHERE 
				MI.docstatus != 2
				{conditions}
			"""
		result = frappe.db.sql(query, as_dict=1)
		for each in result:
			for i in monitoringYears:
				query = f"""
							SELECT
								MCT.actual_monitored_value
							FROM
								`tabMitigation Monitoring Information` AS MI
							INNER JOIN
								`tabMitigation Monitoring Information ChildTable` AS MCT
							ON
								MI.name = MCT.parent
							WHERE
								MI.proj_id = '{project}'
							AND
								MI.monitoring_year = '{i.monitoring_year}'
							AND 
								MCT.performance_indicator = '{each.performance_indicator}'
							"""
				value = frappe.db.sql(query,as_dict =1)

				
				

				each[f'{i.monitoring_year}'] = value[0].actual_monitored_value if value and value[0].actual_monitored_value else 0
		output.extend(result)
		values_only = [list({k: v for k, v in item.items() if k != 'name'}.values()) for item in output]
		return values_only
def get_adaptation_datas(project = None):
	conditions = ""
	if project:
		
		conditions += f" AND AM.project_id like '{project}'"
		query= f"""
			SELECT
				QI.category,
				QI.question,
				QI.expected_value
			FROM
				`tabAdaptation` AS AM
			INNER JOIN
				`tabAdaptation Quantitative ChildTable` as QI
			ON
				AM.name = QI.parent
			WHERE
				AM.docstatus != 2
			AND
				QI.expected_value != 0
				{conditions}
			"""
		result = frappe.db.sql(query, as_dict=1)
		
		monitoringYearsAdaptation = frappe.db.get_all('Adaptation Monitoring Information',
					filters={'proj_id':project},
					fields = ['monitoring_year'],
					order_by = 'monitoring_year asc')
		
		for each in result:
			for i in monitoringYearsAdaptation:
				query = f"""
							SELECT
								ACT.actual_value
							FROM
								`tabAdaptation Monitoring ChildTable` AS ACT
							INNER JOIN
								`tabAdaptation Monitoring Information` AS AMI
							ON
								ACT.parent = AMI.name
							WHERE
								AMI.proj_id = '{project}'
							AND
								AMI.monitoring_year = '{i.monitoring_year}'
							AND 
								ACT.question = '{each.question}'
							"""
				valueAdaptation = frappe.db.sql(query,as_dict =1)
				each[f'{i.monitoring_year}'] = valueAdaptation[0].actual_value if valueAdaptation and valueAdaptation[0].actual_value else 0
		values_only = [list({k: v for k, v in item.items() if k != 'name'}.values()) for item in result]
		return values_only

def get_sdg_datas(project = None):
	conditions = ""
	if project:
		
		conditions += f" AND SDG.project_id like '{project}'"
		query= f"""
			SELECT
				QI.category,
				QI.question,
				QI.data AS expected_value
			FROM
				`tabSDG Assessment` AS SDG
			INNER JOIN
				`tabSDG Quantitative ChildTable` as QI
			ON
				SDG.name = QI.parent
			WHERE
				SDG.docstatus != 2
			AND
				QI.data != 0
				{conditions}
			"""
		result = frappe.db.sql(query, as_dict=1)
		monitoringYearsSDG = frappe.db.get_all('SDG Monitoring Information',
					filters={'proj_id':project},
					fields = ['monitoring_year'],
					order_by = 'monitoring_year asc' )
		for each in result:
			for i in monitoringYearsSDG:
				query = f"""
							SELECT
								ACT.data
							FROM
								`tabSDG Quantitative ChildTable` AS ACT
							INNER JOIN
								`tabSDG Monitoring Information` AS AMI
							ON
								ACT.parent = AMI.name
							WHERE
								AMI.proj_id = '{project}'
							AND
								AMI.monitoring_year = '{i.monitoring_year}'
							AND 
								ACT.question = '{each.question}'
							"""
				valueSDG = frappe.db.sql(query,as_dict =1)
				each[f'{i.monitoring_year}'] = valueSDG[0].data if valueSDG and valueSDG[0].data else 0
		values_only = [list({k: v for k, v in item.items() if k != 'name'}.values()) for item in result]
		return values_only

def get_finance_datas(project = None):
	conditions = ""
	if project:	
		
		amountExpected = {'Finance Data': 'Expected Spend (USD)'}
		amountSpent = {'Finance Data':'Amount Spent (USD)'}
		# name = frappe.db.get_value('Climate Finance', {'project_id': f'{project}'}, 'name')
		# monitoringYearsFinance = frappe.db.get_list(
		# 	"Climate Finance Disbursement Schedule ChildTable",
		# 	fields=['financial_year'],
		# 	filters = {"parent" : name},
		# 	order_by = 'financial_year')
		if frappe.db.exists('Climate Finance',{'project_id': project}):
			expectedDoc = frappe.get_doc('Climate Finance',{'project_id': project})
			expectedDocList = expectedDoc.as_dict().budget_disbursement_schedule
			for i in expectedDocList:
				amountExpected[f'{i.financial_year}'] = i.amount
			
			spentDoc = frappe.get_last_doc('Climate Finance Monitoring Information', {'proj_id': project})
			spentDocList = spentDoc.as_dict().total_budget_disbursement
			for i in spentDocList:
				amountSpent[f'{i.financial_year}']=i.total_disbursement_usd
			
			frappe.log_error("amountExpected",amountExpected)
			frappe.log_error("amountSpent",amountSpent)
			
			return [amountSpent,amountExpected]




@frappe.whitelist(allow_guest=True)
def get_chart(project=None):
	if project:
		get_actual_ghg = []
		expectedValue = frappe.db.get_value('Mitigations',{"project_id":f'{project}'},'expected_annual_ghg')
		monitoringYears = frappe.db.get_all('Mitigation Monitoring Information',
							filters={'proj_id':project},
							fields = ['monitoring_year'],
							order_by = 'monitoring_year asc',
							pluck="monitoring_year")
		
		for i in monitoringYears:
			currentValue = frappe.db.get_value('Mitigation Monitoring Information',{"proj_id":f'{project}',"monitoring_year":i},'actual_annual_ghg')
			# output[0][f'{i.monitoring_year}'] = currentValue if currentValue else 0
			get_actual_ghg.append(currentValue)
		
		return {"data":get_actual_ghg,"labels":monitoringYears}

		# chart = {
		# 	'data': {
		# 	'labels': monitoringYears, 
		# 	'datasets': [{'title': 'GHG Emissions Achieved (tCO2e)', 'values': get_actual_ghg}]
		# 	}, 
		# 	'type': 'bar', 
		# 	'colors': ['pink']
		# }
		# return chart

@frappe.whitelist(allow_guest=True)
def get_chart2(project=None):
	if project:
		year_list=[]
		amountExpected = []
		amountSpent = []
		expectedDoc = frappe.get_doc('Climate Finance',{'project_id': project})
		
		expectedDocList = expectedDoc.as_dict().budget_disbursement_schedule
		
		for i in expectedDocList:
			amountExpected.append(i.amount)
			
		spentDoc = frappe.get_last_doc('Climate Finance Monitoring Information', {'proj_id': project})
		spentDocList = spentDoc.as_dict().total_budget_disbursement
		for i in spentDocList:
			amountSpent.append(i.total_disbursement_usd)
			year_list.append(f"{i.financial_year}")
			


		# name = frappe.db.get_value('Climate Finance', {'project_id': f'{project}'}, 'name')
		# totalMonitoringYearsFinance = frappe.db.get_all(
		# 	"Climate Finance Disbursement Schedule ChildTable",
		# 	fields=['financial_year'],
		# 	filters = {"parent" : name},
		# 	order_by = 'financial_year')
		# for i in totalMonitoringYearsFinance:
		# 	year_list.append(f"{i.financial_year}")
				
		return {"data":amountSpent,"labels":year_list}
		


@frappe.whitelist()

def download_excel(project):
	df1 = get_project_details(project)
	df2 = get_mitigation_details(project)
	df3 =  get_adaptation_details(project)
	df4 =  get_sdg_details(project)
	df5 =  get_finance_details(project)
	
	# frappe.log_error("df1",df1)
	# frappe.log_error("df2",df2)
	# frappe.log_error("df3",df3)
	# frappe.log_error("df4",df4)
	# frappe.log_error("df5",df5)
	
	data_dict1 = {df1[0][i]: [row[i] for row in df1[1]] for i in range(len(df1[0]))}
	data_dict2 = {df2[0][i]: [row[i] for row in df2[1]] for i in range(len(df2[0]))}
	data_dict3 = {df3[0][i]: [row[i] for row in df3[1]] for i in range(len(df3[0]))}
	data_dict4 = {df4[0][i]: [row[i] for row in df4[1]] for i in range(len(df4[0]))}
	data_dict5 = {key: [d.get(key) for d in df5[1]] for key in df5[0]}

	export_data1 = pd.DataFrame(data_dict1)
	export_data2 = pd.DataFrame(data_dict2)
	export_data3 = pd.DataFrame(data_dict3)
	export_data4 = pd.DataFrame(data_dict4)
	export_data5 = pd.DataFrame(data_dict5)
	
	
	site_name = get_site_base_path()
	nowTime = now()[:-7]
	nowTime = nowTime.replace(" ","")
	nowTime = nowTime.replace("-","")
	nowTime = nowTime.replace(":","")
	
	# Initialize the Excel Writer
	wb = Workbook()
	ws = wb.active
	ws.title = 'Sheet1'
	
	# Create a bold font
	bold_font = Font(bold=True)
	
	# Write each dataframe to a different worksheet.
	export_data_list = [export_data1, export_data2, export_data3, export_data4, export_data5]
	for df in range(0,len(export_data_list)):
		rows = dataframe_to_rows(export_data_list[df], index=False, header=True)
		header = next(rows)
		if df == 0:
			project_id = ["Project ID :",project ]
			table_heading = ["Project Details :"]
			ws.append(project_id)
			ws.append([])
			for cell in ws[ws.max_row]:
				cell.font = bold_font
		if df == 1:
			table_heading = ["Mitigation Summary :"]
		if df == 2:
			table_heading = ["Adaptation Summary :"]
		if df == 3:
			table_heading = ["SDG Summary :"]
		if df == 4:
			table_heading = ["Finance Summary :"]


		

		ws.append(table_heading)
		for cell in ws[ws.max_row]:
			cell.font = bold_font

		ws.append(header)
		for cell in ws[ws.max_row]:
			cell.font = bold_font
		
		for row in rows:
			ws.append(row)
		
		ws.append([])
		ws.append([])
		ws.append([])
	
	# Adjust the width of the columns and remove border
	for column_cells in ws.columns:
		length = max(len(str(cell.value)) for cell in column_cells)
		ws.column_dimensions[column_cells[0].column_letter].width = length
		for cell in column_cells:
			cell.border = None
	
	# Save the workbook
	wb.save(f"{site_name}/public/files/MRV-Report-{nowTime}.xlsx")
	return f"../files/MRV-Report-{nowTime}.xlsx"




